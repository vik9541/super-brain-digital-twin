#!/usr/bin/env python3
"""
Reports Generator for Digital Twin System
Generates hourly Excel reports and sends via Email and Telegram
"""

import os
import asyncio
from datetime import datetime, timedelta
from supabase import create_client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.encoders import encode_base64
import telegram


class ReportsGenerator:
    def __init__(self):
        self.supabase = create_client(
            os.getenv("SUPABASE_URL"),
            os.getenv("SUPABASE_KEY")
        )
        self.telegram_bot = telegram.Bot(token=os.getenv("TELEGRAM_BOT_TOKEN"))
        self.smtp_config = {
            "host": os.getenv("SMTP_HOST"),
            "user": os.getenv("SMTP_USER"),
            "password": os.getenv("SMTP_PASSWORD"),
            "from": os.getenv("SMTP_FROM"),
            "to": os.getenv("SMTP_TO").split(",")
        }

    async def run(self):
        """Main report generation function"""
        try:
            # 1. Fetch data
            data = await self.fetch_hourly_data()
            
            # 2. Generate Excel
            excel_file = self.generate_excel_report(data)
            
            # 3. Send email
            await self.send_email_report(excel_file)
            
            # 4. Send Telegram
            await self.send_telegram_report(excel_file)
            
            print(f"Report generated successfully at {datetime.utcnow().isoformat()}")
        except Exception as e:
            print(f"Error generating report: {e}")
            await self.send_error_alert(str(e))

    async def fetch_hourly_data(self) -> list:
        """Fetch data for the last hour"""
        now = datetime.utcnow()
        hour_ago = now - timedelta(hours=1)
        
        response = self.supabase.table("analyses") \
            .select("*") \
            .gte("created_at", hour_ago.isoformat()) \
            .lte("created_at", now.isoformat()) \
            .execute()
        
        return response.data

    def generate_excel_report(self, data: list) -> str:
        """Generate Excel report from data"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hourly Report"
        
        # Headers
        headers = ["ID", "Timestamp", "Status", "Duration (s)", 
                   "Records Processed", "Records Failed", "Success Rate"]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        for row, record in enumerate(data, 2):
            ws.cell(row=row, column=1, value=record.get("id"))
            ws.cell(row=row, column=2, value=record.get("created_at"))
            ws.cell(row=row, column=3, value=record.get("status"))
            ws.cell(row=row, column=4, value=record.get("duration"))
            ws.cell(row=row, column=5, value=record.get("records_processed"))
            ws.cell(row=row, column=6, value=record.get("records_failed"))
            
            # Calculate success rate
            processed = record.get("records_processed", 0)
            failed = record.get("records_failed", 0)
            success_rate = 100 if processed == 0 else ((processed - failed) / processed * 100)
            ws.cell(row=row, column=7, value=f"{success_rate:.1f}%")
        
        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save
        filename = f"/tmp/report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename

    async def send_email_report(self, excel_file: str) -> bool:
        """Send report via email"""
        try:
            msg = MIMEMultipart()
            msg["From"] = self.smtp_config["from"]
            msg["To"] = ", ".join(self.smtp_config["to"])
            msg["Subject"] = f"Digital Twin Hourly Report - {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
            
            # Body
            body = f"""Hourly report for Digital Twin system.

Generated at: {datetime.utcnow().isoformat()}

Please see attached Excel file for details.

Best regards,
Digital Twin Bot
"""
            msg.attach(MIMEText(body, "plain"))
            
            # Attachment
            with open(excel_file, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={excel_file.split('/')[-1]}")
            msg.attach(part)
            
            # Send
            with smtplib.SMTP_SSL(self.smtp_config["host"], 465) as server:
                server.login(self.smtp_config["user"], self.smtp_config["password"])
                server.sendmail(self.smtp_config["from"], self.smtp_config["to"], msg.as_string())
            
            print(f"Email sent to {self.smtp_config['to']}")
            return True
        except Exception as e:
            print(f"Error sending email: {e}")
            return False

    async def send_telegram_report(self, excel_file: str) -> bool:
        """Send report via Telegram"""
        try:
            chat_id = int(os.getenv("TELEGRAM_CHAT_ID"))
            with open(excel_file, "rb") as f:
                await self.telegram_bot.send_document(
                    chat_id=chat_id,
                    document=f,
                    caption=f"Hourly Report - {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
                )
            print(f"Telegram report sent to {chat_id}")
            return True
        except Exception as e:
            print(f"Error sending Telegram report: {e}")
            return False

    async def send_error_alert(self, error: str):
        """Send error alert to Telegram"""
        try:
            chat_id = int(os.getenv("TELEGRAM_CHAT_ID"))
            await self.telegram_bot.send_message(
                chat_id=chat_id,
                text=f"⚠️ ERROR in Reports Generator:\n{error}"
            )
        except:
            pass


if __name__ == "__main__":
    generator = ReportsGenerator()
    asyncio.run(generator.run())
